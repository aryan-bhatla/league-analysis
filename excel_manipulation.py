#----------------------------------------------------------------------------------------------------- # 
# Modules
#----------------------------------------------------------------------------------------------------- # 
import pandas as pd 
import team_analysis

from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


#----------------------------------------------------------------------------------------------------- #
# Fetch odds function
#----------------------------------------------------------------------------------------------------- # 
def fetch_odds(team_a: str, team_b: str, odds_df: pd.DataFrame) -> tuple:
    '''
        
        Parameters: 
            team a, team b (str): Teams read in from Excel file
            odds_df (pd.DataFrame): Odds dataframe containing all teams' odds
            
        Returns: 
            odds_a, odds_b (tuple): Winning odds for teams a and b  

    '''
    try:
        # Fetch odds for Team A as winner and Team B as opponent
        odds_a = odds_df.loc[f"{team_a} (Winner)", f"{team_b} (Opponent)"]
    except KeyError:
        odds_a = None

    try:
        # Fetch odds for Team B as winner and Team A as opponent
        odds_b = odds_df.loc[f"{team_b} (Winner)", f"{team_a} (Opponent)"]
    except KeyError:
        odds_b = None

    return odds_a, odds_b


#----------------------------------------------------------------------------------------------------- #
# Load odds into excel 
#----------------------------------------------------------------------------------------------------- # 
file_path = 'bookmaker_odds.xlsx'
excel_data = pd.read_excel(file_path)

for index, row in excel_data.iterrows():
    team_a = row['Team A '].strip() if pd.notna(row['Team A ']) else None
    team_b = row['Team B '].strip() if pd.notna(row['Team B ']) else None

    if team_a and team_b:
        odds_a, odds_b = fetch_odds(team_a, team_b, team_analysis.best_of_three_lck_odds)
        excel_data.at[index, 'Calculated A'] = odds_a
        excel_data.at[index, 'Calculated B'] = odds_b

updated_file_path = 'updated_bookmaker_odds.xlsx'
excel_data.to_excel(updated_file_path, index=False)


#----------------------------------------------------------------------------------------------------- #
# Highlight cells that provide best odds 
#----------------------------------------------------------------------------------------------------- # 
updated_file_path = 'updated_bookmaker_odds.xlsx'
excel_data = pd.read_excel(updated_file_path)

# Columns containing bookmaker odds for Team A and Team B
bookmaker_columns_a = ['bet365 A', 'neds A', 'Ladbrokes A']
bookmaker_columns_b = ['bet365 B', 'neds B', 'Ladbrokes B']

# Identify the cells to be highlighted
highlight_cells = []

for index, row in excel_data.iterrows():
    # For Team A
    bookmaker_odds_a = [row[col] for col in bookmaker_columns_a if pd.notna(row[col])]
    if bookmaker_odds_a and max(bookmaker_odds_a) > row['Calculated A']:
        max_odd_a = max(bookmaker_odds_a)
        col_name_a = bookmaker_columns_a[bookmaker_odds_a.index(max_odd_a)]
        highlight_cells.append((index, col_name_a))

    # For Team B
    bookmaker_odds_b = [row[col] for col in bookmaker_columns_b if pd.notna(row[col])]
    if bookmaker_odds_b and max(bookmaker_odds_b) > row['Calculated B']:
        max_odd_b = max(bookmaker_odds_b)
        col_name_b = bookmaker_columns_b[bookmaker_odds_b.index(max_odd_b)]
        highlight_cells.append((index, col_name_b))

# Load the workbook and get the active sheet
wb = load_workbook(updated_file_path)
sheet = wb.active

# Define the red fill style
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Apply the red fill to the highlighted cells
for cell in highlight_cells:
    row, col_name = cell
    col_letter = get_column_letter(excel_data.columns.get_loc(col_name) + 1)  # Convert column name to letter
    cell_coordinate = f"{col_letter}{row + 2}"  # Adjust for Excel's 1-based indexing and header row
    sheet[cell_coordinate].fill = red_fill

# Save the styled workbook
wb.save('final_styled_bookmaker_odds.xlsx')